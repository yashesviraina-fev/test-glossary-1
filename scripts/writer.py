from sys import argv

scriptEnable=True;

if(not scriptEnable):
    print("scriptEnable is set to False");
    exit(0);

import openpyxl

mode = 0;

finalData = "";
wb = openpyxl.load_workbook("../Glossary.xlsx");
sheet = wb.worksheets[mode];

headings = [];

for j in range(2,10):
    cell_val = sheet.cell(row = 1,column = j);
    headings.append(cell_val.value);

for i in headings:
    print(i);

for i in range(2,21):
    for j in range(2,10):
        cell_val = sheet.cell(row = i,column = j);
        if(cell_val.value == None or headings[j-2] == "Domain"):
            continue;
        if (j == 2):
            finalData += "## " + cell_val.value if cell_val.value != None else "N/A"
            finalData += "\n";
        elif(j == 6):
            name = (cell_val.value if cell_val.value != None else "N/A");
            if(name != "ZF"):
                finalData += "* " + headings[j-2] + " - <a href=\"" + (cell_val.value if cell_val.value != None else "N/A") + '\">' + (cell_val.value if cell_val.value != None else "N/A") + '</a>';
                finalData += "\n";
            elif(name == "ZF"):
                finalData += "* " + headings[j-2] + " - " + (cell_val.value if cell_val.value != None else "N/A") + "\n";

        else:
            finalData += "* " + headings[j-2] + " - " + (cell_val.value if cell_val.value != None else "N/A");
            finalData += "\n";
    finalData += "\n\n";

file = open("../project/docs/swrefarch_glossary.md" if mode==0 else "../project/docs/classic_glossary.md",'w');
if(file.writable()):
    file.write(finalData);
file.close();
